#!/usr/bin/env python3
"""Apply a set of per-row column updates to an existing .xlsx workbook,
matched by its "place_id" column, leaving every other row/column/format/
data-validation dropdown untouched.

The write-side sibling of xlsx-to-json.py. Kept just as dumb: no business
logic about which fields are allowed, what a valid value looks like, or
what "reviewed" means — that all lives in workbench-sync.mjs, which is the
only caller. This script's only job is: given a set of {place_id: {column:
value}} updates, safely land them in the workbook.

Safety properties (all load-bearing, not incidental):
  - Refuses to run at all if a sibling Excel lock file (~$<filename>) is
    present, rather than risk a corrupt write against an open workbook.
  - Refuses to run at all if ANY place_id in the update set is not found in
    the sheet, rather than silently apply a partial update — a caller
    should never be able to half-write a batch.
  - Refuses to run at all if any updated column name is not an existing
    header — a typo in the caller must be loud, not silently ignored.
  - The workbook is loaded fully (not read_only), modified in memory, and
    saved to a temp file in the same directory; only once that succeeds is
    it atomically renamed over the real path (os.replace, same filesystem)
    — a crash mid-save can never leave the original file corrupt or
    half-written.

Usage: python write-updates-to-xlsx.py <path-to-xlsx>
Input (stdin): {"<place_id>": {"<column name>": <value|null>, ...}, ...}
Output (stdout): {"updated": [<place_id>, ...], "path": "<path>"}
"""
import json
import os
import sys
import tempfile

import openpyxl


def write_json(payload) -> None:
    sys.stdout.buffer.write(json.dumps(payload, ensure_ascii=False).encode("utf-8"))


def fail(message: str) -> int:
    sys.stderr.write(message + "\n")
    return 1


def main() -> int:
    if len(sys.argv) != 2:
        return fail("usage: write-updates-to-xlsx.py <path-to-xlsx> (updates JSON on stdin)")

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        return fail(f"Spreadsheet not found: {xlsx_path}")

    lock_path = os.path.join(os.path.dirname(xlsx_path) or ".", f"~${os.path.basename(xlsx_path)}")
    if os.path.exists(lock_path):
        return fail(
            f"REFUSING TO WRITE: {xlsx_path} appears to be open in Excel "
            f"(lock file present: {lock_path}). Close it and try again."
        )

    try:
        # Read raw bytes and decode explicitly as UTF-8 rather than
        # sys.stdin.read(): on Windows, sys.stdin's default text-mode
        # decoding uses the console/locale codepage (cp1252 here), not
        # UTF-8, which silently mangles any non-ASCII character (confirmed
        # directly — a rupee sign sent as correct UTF-8 bytes came back as
        # three separate cp1252 characters once written into the sheet).
        # Mirrors xlsx-to-json.py's own reasoning for using stdout.buffer
        # instead of sys.stdout, one direction earlier in the pipeline.
        updates = json.loads(sys.stdin.buffer.read().decode("utf-8"))
    except json.JSONDecodeError as err:
        return fail(f"Could not parse updates JSON on stdin: {err}")
    if not isinstance(updates, dict):
        return fail("Updates JSON must be an object of {place_id: {column: value}}.")

    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1))
    headers = {}
    for cell in header_row:
        if cell.value is not None:
            headers[str(cell.value)] = cell.column

    if "place_id" not in headers:
        return fail('Spreadsheet has no "place_id" column — refusing to guess which one it is.')

    # Every column referenced by any update must already exist as a header.
    referenced_columns = {col for fields in updates.values() for col in fields}
    unknown_columns = sorted(referenced_columns - set(headers))
    if unknown_columns:
        return fail(f"Unknown column(s) in updates, not present in the sheet: {', '.join(unknown_columns)}")

    place_id_col = headers["place_id"]
    row_by_place_id = {}
    for row in sheet.iter_rows(min_row=2):
        value = row[place_id_col - 1].value
        if value is not None and str(value) in updates:
            row_by_place_id[str(value)] = row[0].row

    missing = sorted(set(updates) - set(row_by_place_id))
    if missing:
        return fail(
            "REFUSING TO WRITE: these place_id(s) from the update set were not found in the "
            f"sheet — aborting the entire write rather than applying a partial update: {', '.join(missing)}"
        )

    for place_id, fields in updates.items():
        row_number = row_by_place_id[place_id]
        for column_name, value in fields.items():
            # NOT sheet.cell(row=..., column=..., value=value): openpyxl
            # treats value=None there as "no value given" (its sentinel for
            # "just return the cell"), so it silently leaves an existing
            # non-blank cell unchanged instead of clearing it — confirmed
            # directly (a row whose "Number Valid" already held "Bogus"
            # stayed "Bogus" after being "updated" to None this way, while
            # an already-blank cell looked fine purely by coincidence).
            # Assigning .value directly has no such sentinel and genuinely
            # clears the cell.
            sheet.cell(row=row_number, column=headers[column_name]).value = value

    directory = os.path.dirname(xlsx_path) or "."
    fd, tmp_path = tempfile.mkstemp(prefix=".workbench-sync-tmp-", suffix=".xlsx", dir=directory)
    os.close(fd)
    try:
        workbook.save(tmp_path)
        os.replace(tmp_path, xlsx_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise

    write_json({"updated": sorted(updates.keys()), "path": xlsx_path})
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

"""
Beggars Map - Phase 1.12: build the human-review workbook from the
re-verification dataset.

Export/review artifact ONLY. Reads tools/discovery/output/reverification-250.json
and writes an .xlsx. Never touches Supabase, never imports, never modifies the
original discovery dataset, research-state.json, or full-verification-results.json.

Key design point: Claude's research is EVIDENCE, not verification. The workbook
keeps the old Phase 1.11 status beside the new one so shallow calls are visible,
and leaves every Human Decision cell blank for the reviewer to fill.

Usage: python build-reverification-workbook.py
"""

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

HERE = Path(__file__).parent
OUTPUT_DIR = HERE / "output"
SOURCE = OUTPUT_DIR / "reverification-250.json"
DEST = OUTPUT_DIR / "reverification-250-human-review.xlsx"

DECISIONS = ["APPROVE", "REJECT", "UNCERTAIN", "NEEDS MORE RESEARCH"]

# (header, width, wrap)
COLUMNS = [
    ("Review #", 8, False),
    ("Restaurant", 30, True),
    ("Address", 40, True),
    ("Area", 14, False),
    ("Zone", 11, False),
    ("Google Maps URL", 22, False),
    ("Phone", 15, False),
    ("Website", 22, False),
    ("Google Rating", 11, False),
    ("Google Reviews", 11, False),
    ("Claude Previous Status", 20, True),
    ("New Research Status", 20, True),
    ("Status Changed?", 12, False),
    ("Cheap Items <=Rs100", 34, True),
    ("Complete Meal Found", 14, False),
    ("Offering Type", 14, False),
    ("Qualifying Meal", 34, True),
    ("Meal Price", 11, False),
    ("Meal Components", 30, True),
    ("Evidence Source", 60, True),
    ("Evidence URL", 26, False),
    ("Branch Match", 12, False),
    ("Images", 9, False),
    ("Image Sources", 20, True),
    ("Research Date", 13, False),
    ("HUMAN DECISION", 18, False),
    ("Human Verified Meal", 26, True),
    ("Human Verified Price", 14, False),
    ("Human Verified Components", 26, True),
    ("Human Evidence / Source", 26, True),
    ("Human Notes", 30, True),
    ("Verification Date", 14, False),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HUMAN_FILL = PatternFill("solid", fgColor="7030A0")
CHANGED_FILL = PatternFill("solid", fgColor="FFF2CC")
QUALIFY_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LINK_FONT = Font(color="0563C1", underline="single", size=10)


def summarise_cheap_items(items):
    if not items:
        return ""
    return "\n".join(f"• {i['item_name']} — Rs{i['price_rupees']}" for i in items)


def summarise_images(images):
    if not images:
        return ""
    return "\n".join(f"{i.get('image_type', 'OTHER')}: {i.get('source', '')}" for i in images)


def build():
    data = json.loads(SOURCE.read_text(encoding="utf-8"))
    rows = data["results"]

    wb = Workbook()

    # ---------------- Instructions ----------------
    ws_help = wb.active
    ws_help.title = "Instructions"
    ws_help.column_dimensions["A"].width = 118
    help_lines = [
        ("HOW TO USE THIS WORKBOOK", "title"),
        ("", None),
        ("This workbook is for HUMAN REVIEW of the Beggars Map candidate research.", "bold"),
        ("Claude's research is EVIDENCE, not verification. Only YOUR 'APPROVE' decision constitutes human verification.", "bold"),
        ("", None),
        ("WHAT COUNTS AS QUALIFYING", "head"),
        ("APPROVE only if the restaurant has a genuine COMPLETE meal / breakfast / thali / combo / substantial", None),
        ("meal-sized dish for Rs100 or less. Examples that qualify: 2 idli + vada plate, breakfast combo,", None),
        ("poori + sabji, complete tiffin, South/North Indian meals, thali, mess meals, unlimited meals,", None),
        ("rice + dal + sabji + roti, a biryani that is a full meal, a substantial rice bowl, a chicken meal.", None),
        ("Use judgement: a plate of momos or a full ghee pongal plate can be a complete meal. A single item is not.", None),
        ("", None),
        ("DO NOT APPROVE", "head"),
        ("• a single vada, single idli, single roti, single chapati", None),
        ("• tea, coffee", None),
        ("• a snack, a side dish, a small portion, an individual cheap ingredient", None),
        ("• an arbitrary combination of separately priced items", None),
        ("", None),
        ("NEVER build a meal by adding separate item prices together. Idli Rs20 + Vada Rs20 does NOT become a", "bold"),
        ("Rs40 breakfast unless the source itself sells it as a plate / combo / meal.", "bold"),
        ("", None),
        ("BRANCH MATCHING IS MANDATORY", "head"),
        ("Names like 'Udupi Darshini', 'Udupi Upahar', 'New Udupi Grand', 'Madura Darshini' and", None),
        ("'Rajanna Military Hotel' repeat across many unrelated branches in Bengaluru. Evidence must refer to", None),
        ("THIS branch — check address, street, phone, website. Never carry a price over from another branch.", None),
        ("", None),
        ("READING THE STATUS COLUMNS", "head"),
        ("'Claude Previous Status' is the older Phase 1.11 call. 'New Research Status' is this deeper re-verification.", None),
        ("Rows where these differ are highlighted — those are the cases where the first pass was too shallow", None),
        ("(or too generous). Please pay special attention to them.", None),
        ("", None),
        ("Research statuses used: RESEARCHED, EVIDENCE_FOUND, QUALIFIES_PENDING_HUMAN_REVIEW, MATCH_UNCERTAIN,", None),
        ("PRICE_UNKNOWN, PRICE_ABOVE_100, SOURCE_UNAVAILABLE, REJECTED.", None),
        ("QUALIFIES_PENDING_HUMAN_REVIEW means: research found a plausible qualifying meal. It is NOT approval.", "bold"),
        ("HUMAN_VERIFIED_LE_100 is never set by research — only you can create it, by choosing APPROVE here.", "bold"),
        ("", None),
        ("FILLING IN YOUR DECISION", "head"),
        ("1. Pick a value in the purple HUMAN DECISION column: APPROVE / REJECT / UNCERTAIN / NEEDS MORE RESEARCH.", None),
        ("2. If you APPROVE, also fill: Human Verified Meal, Human Verified Price, Human Verified Components,", None),
        ("   Human Evidence / Source, and Verification Date. Add anything useful to Human Notes.", None),
        ("3. Leave the Claude research columns untouched — they are the audit trail.", None),
        ("", None),
        ("The 'Cheap Items <=Rs100' column lists individually cheap items found. These are useful colour for a", None),
        ("listing note, but they do NOT by themselves qualify a restaurant.", None),
    ]
    for line, kind in help_lines:
        cell = ws_help.cell(row=ws_help.max_row + 1 if ws_help.max_row > 1 or ws_help["A1"].value else 1, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if kind == "title":
            cell.font = Font(bold=True, size=14, color="1F3864")
        elif kind == "head":
            cell.font = Font(bold=True, size=11, color="1F3864")
        elif kind == "bold":
            cell.font = Font(bold=True, size=10)
        else:
            cell.font = Font(size=10)

    # ---------------- Review sheet ----------------
    ws = wb.create_sheet("Review")
    for idx, (header, width, _wrap) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=idx, value=header)
        c.fill = HUMAN_FILL if header.startswith("HUMAN") or header.startswith("Human") or header == "Verification Date" else HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 30

    for i, r in enumerate(rows, start=1):
        row = i + 1
        changed = r.get("old_status") != r.get("qualifying_status")
        qualifies = r.get("qualifying_status") == "QUALIFIES_PENDING_HUMAN_REVIEW"
        components = r.get("included_items")
        if isinstance(components, list):
            components = "; ".join(str(x) for x in components)
        values = [
            i,
            r.get("name"),
            r.get("address"),
            r.get("area"),
            r.get("zone"),
            r.get("google_maps_uri") or "",
            r.get("phone") or "",
            r.get("website") or "",
            r.get("google_rating") if r.get("google_rating") is not None else "n/a",
            r.get("google_review_count") if r.get("google_review_count") is not None else "n/a",
            r.get("old_status"),
            r.get("qualifying_status"),
            "YES" if changed else "",
            summarise_cheap_items(r.get("cheap_items_under_100")),
            "YES" if qualifies else "no",
            r.get("offering_type") or "",
            r.get("qualifying_meal_description") or "",
            r.get("price_rupees") if r.get("price_rupees") is not None else "",
            components or "",
            r.get("evidence_description") or "",
            r.get("evidence_source_url") or "",
            r.get("match_confidence") or "",
            len(r.get("image_urls") or []),
            summarise_images(r.get("image_urls")),
            (r.get("price_checked_at") or "")[:10],
            "",  # HUMAN DECISION - deliberately blank
            "", "", "", "", "",  # human fields - blank
        ]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=value)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=COLUMNS[col - 1][2], vertical="top")
            c.font = Font(size=10)
        # clickable URLs
        for col in (6, 8, 21):
            cell = ws.cell(row=row, column=col)
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.font = LINK_FONT
        # rupee formatting
        ws.cell(row=row, column=18).number_format = '"₹"#,##0'
        ws.cell(row=row, column=28).number_format = '"₹"#,##0'
        if changed:
            ws.cell(row=row, column=13).fill = CHANGED_FILL
            ws.cell(row=row, column=11).fill = CHANGED_FILL
            ws.cell(row=row, column=12).fill = CHANGED_FILL
        if qualifies:
            ws.cell(row=row, column=12).fill = QUALIFY_FILL
            ws.cell(row=row, column=15).fill = QUALIFY_FILL

    last_row = len(rows) + 1
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"

    dv = DataValidation(type="list", formula1='"' + ",".join(DECISIONS) + '"', allow_blank=True, showDropDown=False)
    dv.error = "Pick one of: APPROVE, REJECT, UNCERTAIN, NEEDS MORE RESEARCH"
    dv.errorTitle = "Invalid decision"
    dv.prompt = "APPROVE only for a genuine complete meal at Rs100 or less."
    dv.promptTitle = "Human decision"
    ws.add_data_validation(dv)
    dv.add(f"Z2:Z{last_row}")

    decision_range = f"Z2:Z{last_row}"
    ws.conditional_formatting.add(decision_range, CellIsRule(operator="equal", formula=['"APPROVE"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100", bold=True)))
    ws.conditional_formatting.add(decision_range, CellIsRule(operator="equal", formula=['"REJECT"'], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006", bold=True)))
    ws.conditional_formatting.add(decision_range, CellIsRule(operator="equal", formula=['"UNCERTAIN"'], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(color="9C6500", bold=True)))
    ws.conditional_formatting.add(decision_range, CellIsRule(operator="equal", formula=['"NEEDS MORE RESEARCH"'], fill=PatternFill("solid", fgColor="DDEBF7"), font=Font(color="1F4E79", bold=True)))

    # ---------------- Summary ----------------
    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions["A"].width = 46
    ws_sum.column_dimensions["B"].width = 16
    ws_sum.column_dimensions["C"].width = 52

    def put(row, label, value, note="", bold=False, fill=None):
        a = ws_sum.cell(row=row, column=1, value=label)
        b = ws_sum.cell(row=row, column=2, value=value)
        c = ws_sum.cell(row=row, column=3, value=note)
        for cell in (a, b, c):
            cell.font = Font(bold=bold, size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill
        return row + 1

    counts = {}
    old_counts = {}
    for r in rows:
        counts[r["qualifying_status"]] = counts.get(r["qualifying_status"], 0) + 1
        old_counts[r["old_status"]] = old_counts.get(r["old_status"], 0) + 1
    changed_rows = [r for r in rows if r["old_status"] != r["qualifying_status"]]
    with_cheap = [r for r in rows if r.get("cheap_items_under_100")]
    cheap_item_total = sum(len(r.get("cheap_items_under_100") or []) for r in rows)
    with_images = [r for r in rows if r.get("image_urls")]

    r_ = 1
    r_ = put(r_, "BEGGARS MAP — RE-VERIFICATION (PHASE 1.12)", "", "Human review workbook", bold=True)
    r_ = put(r_, "Generated", date.today().isoformat(), "")
    r_ = put(r_, "Source universe (Phase 1.11 records)", data["run"].get("source_universe", 250), "The existing researched candidates")
    r_ = put(r_, "Re-verified so far (rows in this workbook)", len(rows), "")
    r_ += 1
    r_ = put(r_, "NEW RESEARCH STATUS", "", "", bold=True)
    for status in ["QUALIFIES_PENDING_HUMAN_REVIEW", "EVIDENCE_FOUND", "PRICE_UNKNOWN", "PRICE_ABOVE_100", "MATCH_UNCERTAIN", "SOURCE_UNAVAILABLE", "REJECTED", "RESEARCHED"]:
        r_ = put(r_, f"  {status}", counts.get(status, 0), "")
    r_ += 1
    r_ = put(r_, "COMPARISON WITH PHASE 1.11", "", "", bold=True)
    r_ = put(r_, "  Rows whose status CHANGED", len(changed_rows), "The old pass was too shallow or too generous here")
    r_ = put(r_, "  Old SOURCE_UNAVAILABLE", old_counts.get("SOURCE_UNAVAILABLE", 0), "")
    r_ = put(r_, "  New SOURCE_UNAVAILABLE", counts.get("SOURCE_UNAVAILABLE", 0), "Lower is better — evidence was found on re-check")
    r_ = put(r_, "  Old QUALIFIES_PENDING_HUMAN_REVIEW", old_counts.get("QUALIFIES_PENDING_HUMAN_REVIEW", 0), "")
    r_ = put(r_, "  New QUALIFIES_PENDING_HUMAN_REVIEW", counts.get("QUALIFIES_PENDING_HUMAN_REVIEW", 0), "")
    r_ += 1
    r_ = put(r_, "CHEAP ITEMS & IMAGES", "", "", bold=True)
    r_ = put(r_, "  Restaurants with cheap items (<=Rs100) recorded", len(with_cheap), "Useful for listing notes; does not qualify a restaurant")
    r_ = put(r_, "  Total cheap items recorded", cheap_item_total, "")
    r_ = put(r_, "  Restaurants with image evidence", len(with_images), "No image URLs are fabricated; empty where none found")
    r_ += 1
    r_ = put(r_, "HUMAN DECISIONS (live counts)", "", "Updates automatically as you fill the Review sheet", bold=True)
    for label in DECISIONS:
        ws_sum.cell(row=r_, column=1, value=f"  {label}").font = Font(size=10)
        ws_sum.cell(row=r_, column=2, value=f'=COUNTIF(Review!Z:Z,"{label}")').font = Font(size=10, bold=True)
        r_ += 1
    ws_sum.cell(row=r_, column=1, value="  Still blank (not yet reviewed)").font = Font(size=10)
    ws_sum.cell(row=r_, column=2, value=f'={len(rows)}-COUNTA(Review!Z2:Z{last_row})').font = Font(size=10, bold=True)
    r_ += 2
    ws_sum.cell(row=r_, column=1, value="Claude's research is EVIDENCE, not verification. Only an APPROVE decision here is human verification.").font = Font(bold=True, size=10, color="9C0006")

    wb.save(DEST)
    print(f"[workbook] Wrote {DEST}")
    print(f"[workbook] Rows: {len(rows)}  |  status changes: {len(changed_rows)}  |  qualifying: {counts.get('QUALIFIES_PENDING_HUMAN_REVIEW', 0)}")


if __name__ == "__main__":
    build()
